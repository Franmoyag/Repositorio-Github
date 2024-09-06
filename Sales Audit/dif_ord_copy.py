import pymongo
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Mapeo de IDs de sucursales a nombres
sucursal_mapping = {
    1: "Sucursal de prueba",
    2: "Lo Prado",
    3: "Maipú Vespucio",
    4: "Quilicura",
    5: "Santiago Centro",
    6: "Ñuñoa",
    7: "La Florida",
    8: "Independencia",
    9: "Plaza Maipú",
    10: "Renca",
    11: "Restaurant Colombiano",
    12: "San Bernardo",
    13: "Las Condes",
    14: "Puente Alto",
    15: "Lo Espejo",
    16: "NT",
    17: "Patio Egaña",
    18: "El Bosque",
    19: "San Ramón",
    20: "La Cisterna",
    21: "Latin Pizza San Ramon",
    22: "Macul",
    23: "Pedro Aguirre Cerda",
    24: "Estacion Central",
    25: "San Miguel",
    26: "Renca Poniente",
    27: "Buin",
    28: "Conchali",
    29: "San Joaquin",
    30: "Maipú 4 Poniente",
    31: "Peñalolen",
    32: "Las Rejas",
    33: "Huechuraba",
    34: "Peñaflor",
    35: "Colina",
    36: "Cerrillos",
    37: "Marcoleta",
    38: "Puente Alto 2",
    39: "El Salto",
    40: "Talagante",
    41: "Providencia",
    42: "La Granja",
    43: "Colina 2",
    44: "Curacavi",
    45: "Padre Hurtado",
    46: "Viña del Mar",
    47: "Rancagua",
    48: "Villa Alemana",
    49: "Con Con",
    50: "Quillota",
    51: "Latin Pizza Colina",
    52: "Latin Pizza Padre Hurtado",
    53: "Latin Pizza Lo Espejo",
    54: "Latin Pizza Conchali",
    55: "Latin Pizza Las Rejas",
    56: "Latin Pizza Plaza Egaña",
    57: "Latin Pizza Cerrillos",
    58: "Latin Pizza La Granja",
    59: "Latin Pizza Curacavi",
    60: "Latin Pizza Buin",
    61: "Latin Pizza El Bosque",
    62: "Latin Pizza Peñaflor",
    63: "Latin Pizza 4 Poniente",
    64: "Latin Pizza Quinta Normal",
    65: "Latin Pizza Plaza Maipu",
    66: "Latin Pizza El Salto / Recoleta",
    67: "Latin Pizza Santiago",
    68: "Latin Pizza Talagante",
    69: "Latin Pizza Renca",
    70: "Kami Cerrito",
    71: "Kami El Cerro",
    72: "Kami Las Piedras",
    73: "Kami Las Tejas",
    80: "Concepción",
    81: "Concepción 2 Collao",
    82: "San Pedro de la Paz",
    83: "Tomé",
    84: "Coronel",
    85: "Chiguayante",
    86: "Talcahuano",
    87: "Hualpén",
    88: "Lota",
    89: "La Calera",
    90: "Quilpúe",
    91: "Los Ángeles",
    92: "Peñalolen 2",
    93: "Coronel 2",
    94: "San Pablo",
    95: "Con Con (Test)",
    96: "Puente Alto 3",
    97: "La Pintana",
    98: "Machalí",
    99: "Renca Poniente (Test)",
    100: "La Cisterna (Test)",
    101: "Curacavi (Test)",
    1000: "Kami Pruebas",
    2000: "TEST_HB"
}

payment_mapping = {
    0: "EFECTIVO",
    1: "TRANSBANK",
    2: "TRANSFERENCIA",
    3: "SODEXO",
    4: "APLICACIONES",
    5: "PAGO_ONLINE",
    6: "AMIPASS",
    7: "RAPPI"
}


def main():
    client = pymongo.MongoClient("mongodb://localhost:27017/")
    db = client['data_bsale']
    tabla1 = db['orders']
    tabla2 = db['orders_2']
    tabla3 = db['orders_3']
    tabla4 = db['orders_4']

    diferencias_collection = db['diferencias_acumuladas']

    diferencias_collection.delete_many({})  # Limpiar la colección de diferencias acumuladas

    # Comparaciones entre tablas
    comparar_y_almacenar_diferencias(tabla1, tabla2, diferencias_collection, 'orders', 'orders_2')
    comparar_y_almacenar_diferencias(tabla1, tabla3, diferencias_collection, 'orders', 'orders_3')
    comparar_y_almacenar_diferencias(tabla2, tabla3, diferencias_collection, 'orders_2', 'orders_3')
    comparar_y_almacenar_diferencias(tabla1, tabla4, diferencias_collection, 'orders', 'orders_4')
    comparar_y_almacenar_diferencias(tabla2, tabla4, diferencias_collection, 'orders_2', 'orders_4')
    comparar_y_almacenar_diferencias(tabla3, tabla4, diferencias_collection, 'orders_3', 'orders_4')

    exportar_a_excel(diferencias_collection)

    client.close()

def comparar_y_almacenar_diferencias(tabla_origen, tabla_destino, diferencias_collection, nombre_origen, nombre_destino):
    for doc_origen in tabla_origen.find():
        doc_destino = tabla_destino.find_one({"_id": doc_origen["_id"]})

        if doc_destino:
            productos_origen = doc_origen.get('products', [])
            productos_destino = doc_destino.get('products', [])

            indices_origen = {prod.get('index') for prod in productos_origen}
            indices_destino = {prod.get('index') for prod in productos_destino}

            productos_removidos = indices_origen - indices_destino
            productos_agregados = indices_destino - indices_origen

            total_antiguo = sum([prod.get('price', 0) * prod.get('quantity', 1) for prod in productos_origen])
            total_nuevo = sum([prod.get('price', 0) * prod.get('quantity', 1) for prod in productos_destino])

            fecha = doc_origen.get('datetime', '')
            fecha_formateada = ''
            if fecha:
                fecha_formateada = datetime.strptime(fecha, "%Y-%m-%dT%H:%M:%S.%f").strftime("%d-%m-%Y")

            sucursal_id = doc_origen.get('sucursal', '')
            sucursal_nombre = sucursal_mapping.get(sucursal_id, 'Sucursal desconocida')

            payment_id = doc_origen.get('payment', '')
            payment_name = payment_mapping.get(payment_id, 'Pago Desconocido')

            name_client = f"{doc_origen['user'].get('name', '')} {doc_origen['user'].get('apellido', '')}"
            tel_client = doc_origen['user'].get('phone', '')

            extract_origen = doc_origen.get('extract', '')
            extract_destino = doc_destino.get('extract', '')

            for index in productos_removidos:
                producto = next((prod for prod in productos_origen if prod.get('index') == index), None)
                if producto:
                    producto_title = producto.get('title', '')
                    producto_price = producto.get('price', 0)
                    producto_qty = producto.get('quantity', 1)
                    older_amount = producto_price * producto_qty

                    diferencia_documento = {
                        'date': fecha_formateada,
                        'extract_origen': extract_origen,
                        'extract_destino': extract_destino,
                        'store': sucursal_nombre,
                        'working_day': doc_origen.get('jornada', ''),
                        'order': doc_origen.get('index', ''),
                        'product': producto_title,
                        'qty': producto_qty,
                        'client_name': name_client,
                        'tel_client': tel_client,
                        'diff': 'producto_removed',
                        'payment': payment_name,
                        'older_amount': older_amount,
                        'new_amount': '',
                        'old_total': total_antiguo,
                        'new_total': total_nuevo,
                        'diff_total': total_nuevo - total_antiguo,
                        'status': 'Cambio Detectado' if (total_nuevo - total_antiguo) < 0 else ''
                    }

                    # Verificación mejorada de duplicados
                    existing_diff = diferencias_collection.find_one({
                        'order': doc_origen.get('index', ''),
                        'product': producto_title,
                        'qty': producto_qty,
                        'diff': 'producto_removed',
                        'store': sucursal_nombre,
                        'date': fecha_formateada
                    })

                    # Si no hay diferencias anteriores o si hay una nueva modificación (cambio en total)
                    if not existing_diff or existing_diff['new_total'] != total_nuevo:
                        diferencias_collection.insert_one(diferencia_documento)

            for index in productos_agregados:
                producto = next((prod for prod in productos_destino if prod.get('index') == index), None)
                if producto:
                    producto_title = producto.get('title', '')
                    producto_price = producto.get('price', 0)
                    producto_qty = producto.get('quantity', 1)
                    new_amount = producto_price * producto_qty

                    diferencia_documento = {
                        'date': fecha_formateada,
                        'extract_origen': extract_origen,
                        'extract_destino': extract_destino,
                        'store': sucursal_nombre,
                        'working_day': doc_origen.get('jornada', ''),
                        'order': doc_origen.get('index', ''),
                        'product': producto_title,
                        'qty': producto_qty,
                        'client_name': name_client,
                        'tel_client': tel_client,
                        'diff': 'producto_added',
                        'payment': payment_name,
                        'older_amount': '',
                        'new_amount': new_amount,
                        'old_total': total_antiguo,
                        'new_total': total_nuevo,
                        'diff_total': total_nuevo - total_antiguo,
                        'status': 'Cambio Detectado' if (total_nuevo - total_antiguo) < 0 else ''
                    }

                    # Verificación mejorada de duplicados
                    existing_diff = diferencias_collection.find_one({
                        'order': doc_origen.get('index', ''),
                        'product': producto_title,
                        'qty': producto_qty,
                        'diff': 'producto_added',
                        'store': sucursal_nombre,
                        'date': fecha_formateada
                    })

                    # Si no hay diferencias anteriores o si hay una nueva modificación (cambio en total)
                    if not existing_diff or existing_diff['new_total'] != total_nuevo:
                        diferencias_collection.insert_one(diferencia_documento)


def exportar_a_excel(diferencias_collection):
    diferencias_cursor = diferencias_collection.find().sort('date', pymongo.DESCENDING)
    diferencias_df = pd.DataFrame(list(diferencias_cursor))

    if diferencias_df.empty:
        print("No hay diferencias acumuladas en la colección.")
        return
    else:
        diferencias_df['date'] = pd.to_datetime(diferencias_df['date'], format="%d-%m-%Y", errors='coerce')

    diferencias_df['date'] = diferencias_df['date'].dt.strftime("%d-%m-%Y")

    # Asegurarse de que new_amount y older_amount sean numéricos
    diferencias_df['new_amount'] = pd.to_numeric(diferencias_df['new_amount'], errors='coerce').fillna(0)
    diferencias_df['older_amount'] = pd.to_numeric(diferencias_df['older_amount'], errors='coerce').fillna(0)
    
    # Crear columna de diferencia
    #diferencias_df['difference'] = diferencias_df['new_amount'] - diferencias_df['older_amount']

    now = datetime.now()
    mes_actual = now.strftime("%m")
    año_actual = now.strftime("%Y")

    nombre_archivo = f'diferencias-KAMI-{mes_actual}-{año_actual}.xlsx'
    #f'C:/Users/siste/OneDrive/Escritorio/Reportes Python/mozart/sales_audit/diferencias-KAMI-{mes_actual}-{año_actual}.xlsx'

    diferencias_df.to_excel(nombre_archivo, index=False)

    wb = load_workbook(nombre_archivo)
    ws = wb.active

    fill = PatternFill(start_color="FFF5BCA9", end_color="FFF5BCA9", fill_type="solid")
    font = Font(color="FFDF0101")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.column == ws["S1"].column and cell.value == "Cambio Detectado":
                for c in row:
                    c.fill = fill
                    c.font = font

    wb.save(nombre_archivo)
    print(f"Diferencias almacenadas con éxito en MongoDB y exportadas a {nombre_archivo} con estilos aplicados.")

if __name__ == "__main__":
    main()
